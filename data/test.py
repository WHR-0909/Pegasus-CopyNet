import openpyxl  
  
# 打开Excel文件  
workbook = openpyxl.load_workbook('3.xlsx')  
sheet = workbook.active  # 假设要操作的是活动工作表  
  
# 指定要操作的列索引，例如第二列（B列）的索引是2  
column_index = 2  
  
# 获取工作表的最大行数  
max_row = sheet.max_row  
  
# 遍历指定列的所有单元格  
for row in range(1, max_row + 1):  # 从第一行开始遍历到最后一行  
    cell = sheet.cell(row=row, column=column_index)  # 获取指定行和列的单元格  
    cell_value = str(cell.value) if cell.value is not None else ""  
      
    # 检查单元格是否包含"大数据服务"  
    if '技术研发' in cell_value:  
        # 找到"大数据服务"及其后面所有分号的索引  
        start_index = cell_value.find('技术研发')  
        if start_index != -1:  
            end_index = start_index + len('技术研发')  
            while end_index < len(cell_value) and cell_value[end_index] == ';':  
                end_index += 1  
              
            # 构造新的字符串，去掉"大数据服务"及其后面的所有分号  
            new_value = cell_value[:start_index] + cell_value[end_index:]  
              
            # 更新单元格的值  
            cell.value = new_value  
  
# 保存修改后的Excel文件  
workbook.save('your_excel_file_modified.xlsx')
